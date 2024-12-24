import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from mailing import letter, parcel, package
from docx import Document
from openpyxl import Workbook

def calculate():
    t = combobox.get()
    distance = float(distance_entry.get())
    time = float(time_entry.get())
    weight = float(weight_entry.get())
    volume = float(volume_entry.get())
    delivery_type = delivery_type_entry.get()

    if t == "letter":
        price = letter.calculate_price(distance, time, weight)
    elif t == "parcel":
        price = parcel.calculate_price(distance, time, weight, volume)
    elif t == "package":
        price = package.calculate_price(distance, time, weight, volume, delivery_type)

    label_res.config(text=f"price: {price}")
    global last_res
    last_res = {"type":t,"price":price}

def save():
    filetypes = [("Word Document", "*.docx"), ("Excel Workbook", "*.xlsx")]
    file = filedialog.asksaveasfilename(filetypes=filetypes)

    if file.endswith(".docx"):
        doc = Document()
        for key, value in last_res.items():
            doc.add_paragraph(f"{key}: {value}")
        doc.save(file)
    elif file.endswith(".xlsx"):
        wb = Workbook()
        ws = wb.active
        row = 1
        for key, value in last_res.items():
            ws.cell(row=row, column=1, value=key)
            ws.cell(row=row, column=2, value=value)
            row+=1
        wb.save(file)
    
root = tk.Tk()
root.title("DOOM ETERNAL")

tk.Label(root, text="what do u want to deliver:").grid(row=0, column=0)
combobox = ttk.Combobox(root, values=["letter","parcel","package"])
combobox.grid(row=0,column=1)

tk.Label(root, text="distance").grid(row=1, column=0)
distance_entry = tk.Entry(root)
distance_entry.grid(row=1,column=1)

tk.Label(root, text="time").grid(row=2, column=0)
time_entry = tk.Entry(root)
time_entry.grid(row=2,column=1)

tk.Label(root, text="weight").grid(row=3, column=0)
weight_entry = tk.Entry(root)
weight_entry.grid(row=3,column=1)

tk.Label(root, text="volume").grid(row=4, column=0)
volume_entry = tk.Entry(root)
volume_entry.grid(row=4,column=1)

tk.Label(root, text="delivery type (express?)").grid(row=5, column=0)
delivery_type_entry = tk.Entry(root)
delivery_type_entry.grid(row=5,column=1)

tk.Button(root, text="calculate", command=calculate).grid(row=6,column=0)
tk.Button(root, text="save report", command=save).grid(row=6,column=1)

label_res = tk.Label(root, text="price: ")
label_res.grid(row=7, column=0, columnspan=2)
last_res = {}

root.mainloop()